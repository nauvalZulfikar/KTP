"""API endpoints for reading scheduling scenarios from Excel sheets."""

from __future__ import annotations

from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException, status
from pydantic import BaseModel

from app.config import settings
from scripts.import_excel import REQUIRED_COLUMNS

router = APIRouter(prefix="/scenarios", tags=["scenarios"])

# Sheets to skip (metric sheets, not scenario data)
SKIP_SHEETS = {"Machine Util", "Product WT", "Component WT", "Late Orders"}


def _safe_float(value: Any) -> float | None:
    if not pd.notna(value):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_str(value: Any) -> str | None:
    if not pd.notna(value):
        return None
    text = str(value).strip()
    return text or None


class ScenarioTask(BaseModel):
    unique_id: int
    sr_no: int | None = None
    product_name: str
    order_processing_date: datetime
    promised_delivery_date: datetime
    quantity_required: int
    component: str
    operation: str | None = None
    process_type: str | None = None
    machine_number: str
    run_time_per_1000: float
    cycle_time_seconds: float | None = None
    setup_time_seconds: float | None = None
    start_time: datetime | None = None
    end_time: datetime | None = None


class ScenarioRead(BaseModel):
    name: str
    tasks: list[ScenarioTask]


def _read_scenario_sheet(sheet_name: str) -> list[ScenarioTask]:
    df = pd.read_excel(settings.excel_path, sheet_name=sheet_name)
    # Skip header rows that are just titles (check for required columns)
    if "UniqueID" not in df.columns:
        # Try reading from row 2 (skip title row)
        df = pd.read_excel(settings.excel_path, sheet_name=sheet_name, skiprows=2)
    if "UniqueID" not in df.columns:
        return []
    df = df[df["UniqueID"].notna()].reset_index(drop=True)
    tasks: list[ScenarioTask] = []
    for _, row in df.iterrows():
        data: dict[str, Any] = {}
        data["unique_id"] = int(row["UniqueID"])
        data["sr_no"] = int(row["Sr. No"]) if pd.notna(row.get("Sr. No")) else None
        data["product_name"] = str(row["Product Name"])
        data["order_processing_date"] = pd.Timestamp(row["Order Processing Date"]).to_pydatetime()
        data["promised_delivery_date"] = pd.Timestamp(row["Promised Delivery Date"]).to_pydatetime()
        data["quantity_required"] = int(row["Quantity Required"])
        data["component"] = str(row["Components"])
        data["operation"] = _safe_str(row.get("Operation"))
        data["process_type"] = _safe_str(row.get("Process Type"))
        data["machine_number"] = str(row["Machine Number"])
        run_time = _safe_float(row.get("Run Time (min/1000)"))
        if run_time is None:
            continue
        data["run_time_per_1000"] = run_time
        data["cycle_time_seconds"] = _safe_float(row.get("Cycle Time (seconds)"))
        data["setup_time_seconds"] = _safe_float(row.get("Setup time (seconds)"))
        if "Start Time" in df.columns and pd.notna(row.get("Start Time")):
            data["start_time"] = pd.Timestamp(row["Start Time"]).to_pydatetime()
        if "End Time" in df.columns and pd.notna(row.get("End Time")):
            data["end_time"] = pd.Timestamp(row["End Time"]).to_pydatetime()
        tasks.append(ScenarioTask(**data))
    return tasks


@router.get("", response_model=list[str])
def list_scenarios() -> list[str]:
    """Return names of scenario sheets in the Excel file."""
    wb = openpyxl.load_workbook(settings.excel_path, read_only=True)
    names = [n for n in wb.sheetnames if n not in SKIP_SHEETS]
    wb.close()
    return names


@router.get("/{name}", response_model=ScenarioRead)
def get_scenario(name: str) -> ScenarioRead:
    """Read a specific scenario sheet from the Excel file."""
    wb = openpyxl.load_workbook(settings.excel_path, read_only=True)
    if name not in wb.sheetnames:
        wb.close()
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Sheet '{name}' not found")
    wb.close()
    tasks = _read_scenario_sheet(name)
    return ScenarioRead(name=name, tasks=tasks)
